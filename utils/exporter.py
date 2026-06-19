"""
utils/exporter.py — Export Excel formattato dell'analisi
"""
from io import BytesIO
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# Palette colori BSC
C_HEADER    = "1A5276"   # blu scuro
C_SUBTOTAL  = "D6EAF8"   # azzurro chiaro
C_TOTAL     = "1A5276"   # blu scuro
C_POS       = "E9F7EF"   # verde chiaro
C_NEG       = "FDEDEC"   # rosso chiaro
C_WHITE     = "FFFFFF"
C_GRAY      = "F2F3F4"


def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style(ws, row, col, text, width=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=C_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin_border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _num_format(cell, value, is_pct=False):
    cell.value = value
    cell.number_format = '#,##0.00%' if is_pct else '#,##0.00'
    cell.alignment = Alignment(horizontal="right")
    cell.border = _thin_border()


def _write_ce_section(ws, rows, start_row, section_title, col_value=3):
    """Scrive una sezione CE (righe di tipo item/subtotal/total)."""
    r = start_row
    # Titolo sezione
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_value)
    c = ws.cell(row=r, column=1, value=section_title)
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.fill = PatternFill("solid", fgColor=C_HEADER)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1

    for item in rows:
        row_type = item.get("type", "item")
        label    = item["label"]
        value    = item.get("value", 0.0)

        if row_type == "item":
            ws.cell(row=r, column=1, value="    " + label).border = _thin_border()
            ws.cell(row=r, column=2, value=value)
            ws.cell(row=r, column=2).number_format = '#,##0.00'
            ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=2).border = _thin_border()
            ws.row_dimensions[r].height = 16

        elif row_type == "subtotal":
            for col in range(1, col_value + 1):
                cell = ws.cell(row=r, column=col)
                cell.fill = PatternFill("solid", fgColor=C_SUBTOTAL)
                cell.font = Font(bold=True, size=10)
                cell.border = _thin_border()
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=value)
            ws.cell(row=r, column=2).number_format = '#,##0.00'
            ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
            ws.row_dimensions[r].height = 18

        elif row_type == "total":
            for col in range(1, col_value + 1):
                cell = ws.cell(row=r, column=col)
                cell.fill = PatternFill("solid", fgColor=C_HEADER)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.border = _thin_border()
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=value)
            ws.cell(row=r, column=2).number_format = '#,##0.00'
            ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=2).font = Font(bold=True, color="FFFFFF")
            ws.row_dimensions[r].height = 20

        r += 1

    return r + 1


def export_analysis_excel(
    client_name: str,
    period: str,
    ce_va_rows: list[dict],
    ce_cv_rows: list[dict],
    sp_fin: dict,          # {"attivo": [...], "passivo": [...]}
    sp_fun_rows: list[dict],
    indices: list[dict],
) -> bytes:
    """
    Genera un file Excel con 5 fogli:
    1. CE Valore Aggiunto
    2. CE Costo del Venduto
    3. SP Finanziario
    4. SP Funzionale
    5. Indici
    """
    wb = openpyxl.Workbook()

    # ── 1. CE Valore Aggiunto ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "CE Valore Aggiunto"
    ws1.column_dimensions["A"].width = 48
    ws1.column_dimensions["B"].width = 18
    _write_title(ws1, client_name, period, "Conto Economico — Riclassificazione a Valore Aggiunto")
    _write_ce_section(ws1, ce_va_rows, 4, "CONTO ECONOMICO A VALORE AGGIUNTO")

    # ── 2. CE Costo del Venduto ─────────────────────────────────────────────
    ws2 = wb.create_sheet("CE Costo del Venduto")
    ws2.column_dimensions["A"].width = 48
    ws2.column_dimensions["B"].width = 18
    _write_title(ws2, client_name, period, "Conto Economico — Riclassificazione a Costo del Venduto")
    _write_ce_section(ws2, ce_cv_rows, 4, "CONTO ECONOMICO A COSTO DEL VENDUTO")

    # ── 3. SP Finanziario ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("SP Finanziario")
    ws3.column_dimensions["A"].width = 44
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 18
    _write_title(ws3, client_name, period, "Stato Patrimoniale — Schema Finanziario")
    r = 4
    r = _write_ce_section(ws3, sp_fin.get("attivo", []), r, "ATTIVO")
    _write_ce_section(ws3, sp_fin.get("passivo", []), r, "PASSIVO E PATRIMONIO NETTO")

    # ── 4. SP Funzionale ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("SP Funzionale")
    ws4.column_dimensions["A"].width = 44
    ws4.column_dimensions["B"].width = 18
    _write_title(ws4, client_name, period, "Stato Patrimoniale — Schema Funzionale (CIN = PFN + PN)")
    _write_ce_section(ws4, sp_fun_rows, 4, "SP FUNZIONALE")

    # ── 5. Indici ───────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Indici")
    ws5.column_dimensions["A"].width = 12
    ws5.column_dimensions["B"].width = 42
    ws5.column_dimensions["C"].width = 16
    ws5.column_dimensions["D"].width = 18
    _write_title(ws5, client_name, period, "Indici Economico-Finanziari")
    _write_indices(ws5, indices, start_row=4)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_title(ws, client_name, period, section):
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"BSC Management — {client_name} — {period}"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_HEADER)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:D2")
    c2 = ws["A2"]
    c2.value = section
    c2.font = Font(italic=True, size=10, color="555555")
    c2.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18


def _write_indices(ws, indices, start_row):
    STATUS_EMOJI = {"green": "✅", "yellow": "⚠️", "red": "🔴", "info": "ℹ️"}

    # Header
    headers = ["Gruppo", "Indicatore", "Valore", "Valutazione"]
    for col, h in enumerate(headers, 1):
        _header_style(ws, start_row, col, h)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    prev_group = None
    for idx in indices:
        group = idx["group"]
        if group != prev_group:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            gc = ws.cell(row=r, column=1, value=f"━━  {group}  ━━")
            gc.font = Font(bold=True, color="FFFFFF")
            gc.fill = PatternFill("solid", fgColor="1A5276")
            gc.alignment = Alignment(horizontal="left")
            ws.row_dimensions[r].height = 18
            r += 1
            prev_group = group

        ws.cell(row=r, column=1, value="")
        ws.cell(row=r, column=2, value=idx["label"]).border = _thin_border()
        val_cell = ws.cell(row=r, column=3, value=idx["formatted"])
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.border = _thin_border()
        status_cell = ws.cell(row=r, column=4,
                               value=STATUS_EMOJI.get(idx["status"], "") + " " +
                               {"green":"Buono","yellow":"Attenzione","red":"Critico","info":"Info"}.get(idx["status"],""))
        status_cell.border = _thin_border()

        # Background per status
        if idx["status"] == "green":
            val_cell.fill = PatternFill("solid", fgColor="D5F5E3")
        elif idx["status"] == "yellow":
            val_cell.fill = PatternFill("solid", fgColor="FDEBD0")
        elif idx["status"] == "red":
            val_cell.fill = PatternFill("solid", fgColor="FADBD8")

        ws.row_dimensions[r].height = 16
        r += 1
